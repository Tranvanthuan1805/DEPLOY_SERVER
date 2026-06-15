import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def check_xlsx():
    try:
        wb = openpyxl.load_workbook('c:/Users/duotech/OneDrive/Desktop/DONE_WEB_EDUPATH_AI/thpt_exam_db_import_template.xlsx')
        print("Sheets in workbook:", wb.sheetnames)
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"\nSheet: {name}, dimensions: {sheet.dimensions}")
            # Print first 5 rows
            for row in list(sheet.iter_rows(values_only=True))[:5]:
                print(row)
    except Exception as e:
        print("Error reading Excel file:", e)

check_xlsx()
